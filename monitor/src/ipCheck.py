#! /usr/bin/python

# Oracle Libraries
import cx_Oracle

# Libraries used for Statistical functions
import pandas as pd

# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import datetime

def printRow (row, font, sheet, max_row):
    
    for i in range(0, len(row)):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font 
        sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00' 
    return;

    
def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "ISBillingOperations@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("ISBillingOperations@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def perDif (dval, iq0, iq3):
    
    pdl = []
    
    for x in dval:

        if x > iq3:
            pdl.append(((x - iq3) / iq3) * 100)
        elif x < iq0:
            pdl.append(((x - iq0) / iq0) * 100)
        else:
            pdl.append(0.00)

    return pdl;


def analyse (aip, ahome, aroam, aipDate):
    
    row = []
    
    try:
        bl = [''] * 7
    
        home_desc = list(((pd.DataFrame(ahome)).describe())[0])
        roam_desc = list(((pd.DataFrame(aroam)).describe())[0])
    
        home_last_min = min(ahome[-5:])
        home_last_max = max(ahome[-5:])
    
        aroam_last_min = min(aroam[-5:])
        aroam_last_max = max(aroam[-5:])
    
        if (home_desc[7] > 1000 or roam_desc[7] > 1000) and ((home_last_min < home_desc[4] 
            or roam_last_min < roam_desc[4]) or (home_last_max > home_desc[6] or roam_last_max > roam_desc[6])) :
        
            row.append((aip, float(home_desc[5]), float(home_desc[4]), float(home_desc[6]), float(home_desc[1]), float(home_desc[2]), float(home_desc[7]), float(home_desc[3]),
                float(roam_desc[5]), float(roam_desc[4]), float(roam_desc[6]), float(roam_desc[1]), float(roam_desc[2]), float(roam_desc[7]), float(roam_desc[3])))
            
            homeDif = perDif(ahome[-5:], home_desc[4], home_desc[6])
            roamDif = perDif(aroam[-5:], roam_desc[4], roam_desc[6])
            
            row = row + list(zip(aipDate[-5:], ahome[-5:], homeDif, bl, bl, bl, bl, bl, aroam[-5:], roamDif, bl, bl, bl, bl))
                 
    except:pass
    
    return row;


def dbConnect ():
    
#     CONN_INFO = {
#         'host': '10.176.199.19',  # info from tnsnames.ora
#         'port': 1530,  # info from tnsnames.ora
#         'user': 'md1dbal1',
#         'psw': '{password}',
#         'service': 'bodsprd_adhoc'  # info from tnsnames.ora
# }
#     CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)
#  
#     tconn = cx_Oracle.connect(CONN_STR)

    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_SVC_BILLINGOPS")
    
    return tconn;

###### Main Program  


timeStamp = (datetime.now()).strftime("%Y%m%d")

# Setup Excel 
excel_file = "IPCheck_" + timeStamp + '.xlsx'
wb = Workbook()

sheet = wb.active 
sheet.title = "Oddities"

# Report Headings
header1 = ['IP Address', 'Home Medium', 'IQ0', 'IQ3', 'Home Mean', 'Home STD', 'Home Max', 'Home Min', 'Roam Medium', 'IQ0', 'IQ3', 'Roam Mean', 'Roam STD', 'Roam Max', 'Roam Min']
header2 = ['Date', 'Home Total Records', 'Percent Difference', '', '', '', '', '', 'Roam Total Records', 'Percent Difference']

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=True)

# DB  
conn = dbConnect()
cursor = conn.cursor() 

sql = """ select to_char(start_time,'YYYYMMDD'), l9_ip_address, L9_NT_ROAMING_IND, count(*) as RECORDS, SUM(L3_VOLUME) as VOLUME 
From ape1_rated_event where L3_PAYMENT_CATEGORY = 'PRE' and l3_call_source in ('L','D') 
and (start_time >= (sysdate - 31)) and (to_char(start_time,'YYYYMMDD') <  to_char(sysdate,'YYYYMMDD')) 
group by to_char(start_time,'YYYYMMDD'),l9_ip_address, L9_NT_ROAMING_IND order by 2,1,3"""   

cursor.execute(sql)

results = []
# with open("/home/dbalchen/workspace/monitor/src/ipTest.csv", "rb") as fp:
#     for i in fp.readlines():
#         tmp = i.split("\t")
#         try:
#             results.append((tmp[0], tmp[1], tmp[2], float(tmp[3]), float(tmp[4])))
#         except:pass
results = cursor.fetchall()

for ip_number in sorted(set(map(lambda x:x[1], results))):
    ipList = [x for x in results if x[1] == ip_number]
    
    home = []
    roam = []
    
    all_ip_dates = sorted(set(map(lambda x:x[0], ipList)))
    
    for ip_date in all_ip_dates:
        ipListDate = [x for x in ipList if x[0] == ip_date]
        
        data = ([x for x in ipListDate if x[2] == 'Y'])
        
        if len(data) > 0:
            roam.append((data[0])[3])
        else :
            roam.append(float(0))  
          
        data = ([x for x in ipListDate if x[2] == 'N'])
        
        if len(data) > 0:
            home.append((data[0])[3])
        else :
          home.append(float(0))    
    
    output = analyse(ip_number, home, roam, all_ip_dates)

    curFont = bold_font
    
    if len(output) > 0:
        max_row = sheet.max_row + 1
        
        if max_row == 2:
           printRow(header1, curFont, sheet, 1) 
           
           printRow(output.pop(0), def_font, sheet, 2)
           output.insert(0, header2)
           
           max_row = 4
        
        for row in output:
            
            if (type(row[2]) != str) and ((abs(row[2]) >= 30.0 or abs(row[9]) >= 30.0)) and not row[7]:
                curFont = red_font
                
            printRow(row, curFont, sheet, max_row)

            curFont = def_font                   
            max_row = max_row + 1  
                  
        sheet.append([""])
        
cursor.close

conn.close()

wb.save(excel_file)

message = "Attached is the Pre-pay network data anomalies for " + timeStamp + ". Rows in red show either a new maximum or minimum"
subject = " IP Usage Check for " + timeStamp
sendTo = ["david.balchen@uscellular.com", "ISBillingOperations@uscellular.com"]

for who in sendTo:
    sendMail(excel_file, message, subject, who)

SystemExit(0);

