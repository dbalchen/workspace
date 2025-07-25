#! /usr/bin/python3

# import Workbook used for writing

from openpyxl import Workbook
from openpyxl.styles import Font
# from openpyxl.drawing.image import Image

import fileinput

import argparse

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=True)


def printRow (row, font, sheet, max_row):
    
    for i in range(0, len(row)):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font 
        sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00' 
    return;


def printSheet (sheet, output):
    max_row = 1
    myfont = bold_font
    
    for row in output:
        if(max_row > 1) :
            myfont = def_font
            
        printRow(row, myfont, sheet, max_row)
        
        max_row = max_row + 1 
    return


parser = argparse.ArgumentParser()

parser.add_argument("--input", "-i", help="Input CSV Files")

parser.add_argument("--output", "-o", help="Output File Name")

parser.add_argument("--tabnames", "-t", help="Tab Names")

args = parser.parse_args()

text = ""

lineCount = 0

if args.input:
    inp = args.input
else:
    inp = '-'

# inp = '/home/dbalchen/workspace/usageUtlities/work/sheetout.dat'

for line in fileinput.input(inp):
    try:
        text = text + line
        lineCount = lineCount + 1
    except:pass

if lineCount <= 1:
    exit(0);
    
inP = []
inP = text.split("::")

excel_file = 'statSheet.xlsx'

if args.output:
    excel_file = args.output

sheetList = []
if args.tabnames:
    sheetList = args.tabnames.split("::")
    
# Setup First Sheet
wb = Workbook()

sheet = None

for file in inP: 
        
    results = []
    
    sheetname = "sheet1"
    
    if len(sheetList) > 0:
        sheetname = sheetList[0]
        sheetList.pop(0)
        
    if sheet == None:
        sheet = wb.active
        sheet.title = sheetname
    else :
        sheet = wb.create_sheet(sheetname)
    
#     img = Image('test.png')
#     img.anchor = 'AF1'
#     img.width = img.width*0.7
#     img.height = img.height*0.7
#     sheet.add_image(img)
    
    try:
        for line in fileinput.input(file):
                line = line.rstrip()
                results.append(tuple(line.split("\t")))
    except:
        allLines = []
        allLines = file.split("\n")
        
        for line in allLines:
                line = line.rstrip()
                results.append(tuple(line.split("\t")))
        
    printSheet (sheet, results)   
wb.save(excel_file)

SystemExit(0);

