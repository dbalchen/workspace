#! /usr/bin/python

# import Workbook used for writing

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image

import fileinput

import argparse


# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=True)

def printRow (row, font, sheet, max_row):
    
    for i in range(0, len(row)):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font 
        sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00' 
    return;

def printSheet (sheet, output):
    max_row = 1
    myfont = bold_font
    
    for row in output:
        if(max_row > 1) :
            myfont = def_font
            
        printRow(row,myfont, sheet, max_row)
        
        max_row = max_row + 1 
    return


parser = argparse.ArgumentParser()

parser.add_argument("--input", "-i", help="Input CSV Files")

parser.add_argument("--output", "-o", help="Output File Name")

args = parser.parse_args()

inp = '-'
if args.input:
    inp = args.input
inp = inp.split(";")

excel_file = 'statSheet.xlsx'

if args.output:
    excel_file = args.output

# Setup First Sheet
wb = Workbook()

sheet = None

for file in inp: 
            
    sheetname = (file.split('/'))[-1]   
     
    results = []
    
    if file == '-':
        sheetname = "sheet1"
 
    if sheet == None:
        sheet = wb.active
        sheet.title = sheetname
    else :
        sheet = wb.create_sheet(sheetname)
    
    img = Image('test.png')
    img.anchor = 'AF1'
    img.width = img.width*0.7
    img.height = img.height*0.7
    sheet.add_image(img)
    
    for line in fileinput.input(file):
        try:
            line = line.rstrip()
            results.append(tuple(line.split("\t")))
        except:pass
    printSheet (sheet,results)   
wb.save(excel_file)

SystemExit(0);

