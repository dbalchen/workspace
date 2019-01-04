#! /usr/bin/python

# import Workbook used for writing
from openpyxl import Workbook

# import load_workbook used for reading
from openpyxl import load_workbook

# set file path
filepath="demo.xlsx"

# create Workbook object
wb=Workbook()

# save workbook 
wb.save(filepath)

# Load Workbook
wb=load_workbook(filepath)

# select sheet from demo.xlsx
sheet=wb.active


# Several different ways to add data here are some examples

# set value for cell A1=1
sheet['A1'] = 1

# set value for cell B2=2
sheet.cell(row=2, column=2).value = 2

# Append a group of values to the bottom of the spreadsheet
data=[('Id','Name','Marks'),
      (1,"ABC",50),
      (2,"CDE",100)]

# append all rows
for row in data:
    sheet.append(row)

# save workbook 
wb.save(filepath)


# Let's Open the spreadsheet and read from it...

# load demo.xlsx 
wb=load_workbook(filepath)

# select demo.xlsx
sheet=wb.active

# get b1 cell value
b3=sheet['B3'].value

# get b2 cell value
b4=sheet['B4'].value

# get b3 cell value
b5=sheet.cell(row=5,column=2).value

# print b3, b4 and b5
print(b3)
print(b4)
print(b5)


# Iterating by rows...

# get max row count
max_row=sheet.max_row
# get max column count
max_column=sheet.max_column
# iterate over all cells 
# iterate over all rows
for i in range(1,max_row+1):
     
# iterate over all columns
    for j in range(1,max_column+1):
        # get particular cell value    
        cell_obj=sheet.cell(row=i,column=j)
        # print cell value     
        print(cell_obj.value)
        # print new line
        print('\n')




# Good Bye
SystemExit(0);
