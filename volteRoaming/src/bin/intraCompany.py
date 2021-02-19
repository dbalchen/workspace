#! /bin/python3

# Oracle Libraries
import cx_Oracle

import fileinput

# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from dateutil.relativedelta import relativedelta
from datetime import datetime

import sys

######## Argument Date SQL Setup

date = "20201201" #sys.argv[1]

day = date[6:8]
month = date[4:6]
monthName = (datetime.strptime(date, '%Y%m%d')).strftime("%B")
year = date[0:4]

bp_start_date_LTE = year + month + '01'

title = """Inter-Company late Usage Report - Billing Period """ + monthName + """ """ + day + """ """ + year 
mess = " Attached to this email is the Inter-Company late Usage Report for " + monthName + """ """ + day + """ """ + year + ".\n The report covers the Billing Period Start Date of " + bp_start_date_LTE

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)

sqlDictionary = {}

sqlDictionary["IR"] = """
SELECT To_char (t1.sys_creation_date, 'YYYY-MM-DD')"Creation Date", 
           t1.nr_param_3_val                           "Company Code", 
           t1.carrier_cd "Serving Company",
           rate_plan_cd "Rate Plan",
           QUAL_PARAM_1_VAL "Usage Type",
           ORIG_BP,
           BP_START_DATE,
           count(*) "Record Totals",
           CASE 
             WHEN t1.rate_plan_cd not like '%DATA%' 
                  THEN Sum(t1.tot_chrg_param_val) 
             WHEN t1.rate_plan_cd like '%DATA%'  
                  AND t1.uom = 'K' THEN Sum ( t1.tot_chrg_param_val / 1024 ) 
             ELSE Sum (( t1.tot_chrg_param_val / 1024 ) / 1024) 
           END                                         AS Total_usage, 
           Sum (t1.tot_net_usage_chrg)                 "Total Charges" 
    FROM   ic_accumulated_usage t1 
    WHERE  t1.prod_cat_id = 'IR'
    AND t1.BP_START_DATE = TO_DATE ('""" + bp_start_date_LTE + """', 'YYYYMMDD')
    and ORIG_BP < BP_START_DATE
    GROUP BY TO_CHAR (t1.sys_creation_date, 'YYYY-MM-DD'),
             t1.nr_param_3_val,
             t1.carrier_cd,
             rate_plan_cd,
             QUAL_PARAM_1_VAL,
             ORIG_BP,
             BP_START_DATE,
             t1.uom
    ORDER BY TO_CHAR (t1.sys_creation_date, 'YYYY-MM-DD'),t1.nr_param_3_val 
"""

def printRow (row, font, sheet, max_row):
    
    for i in range(0, len(row)):
        
        if i == 5 or i == 6:
            sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
            sheet.cell(row=max_row, column=(i + 1)).font = font 
        else:
            sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
            sheet.cell(row=max_row, column=(i + 1)).font = font 
            sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00'
    return;

 
def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "david.balchen@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("david.balchen@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def dbConnect ():
    
#     CONN_INFO = {
#         'host': '10.176.199.19',  # info from tnsnames.ora
#         'port': 1530,  # info from tnsnames.ora
#         'user': 'md1dbal1',
#         'psw': 'Potat000#',
#         'service': 'bodsprd_adhoc'  # info from tnsnames.ora
#     }
#     CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)
#   
#     tconn = cx_Oracle.connect(CONN_STR)

    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_SVC_BILLINGOPS")
    
    return tconn;



def printSheet (title, header, sheet, output):
    sheet.title = title
    
    printRow(header, bold_font, sheet, 1)
    max_row = 2 

    for row in output:
        printRow(row, def_font, sheet, max_row)
        max_row = max_row + 1 
    return

###### Main Program  

excel_file = title + '.xlsx'
wb = Workbook()


results = []

# for line in fileinput.input("/home/dbalchen/Desktop/lateU.csv"):
#     try:
#         line = line.rstrip()
#         results.append(tuple(line.split("\t")))
#     except:pass
    
conn = dbConnect()
cursor = conn.cursor()
 
results = []
 
print("Retrieve IR data from Database\n")
 
cursor.execute(sqlDictionary["IR"])
                
results = cursor.fetchall()

printSheet('Inter-company Report', ['Creation Date', 'Company Code', 'Rate Plan', 'Usage Type', 'Original BP', 'BP Start Date', 'Total Records', 'Total Usage', 'Total Charges'], wb.active, results)
#     
# cursor.close
# conn.close()

wb.save(excel_file)

message = mess
subject = title
sendTo = ["david.balchen@uscellular.com"]
#sendTo = ["david.balchen@uscellular.com", 'Philip.Luzod@uscellular.com', 'ISBillingOperations@uscellular.com', 'Ilham.Elgarni@uscellular.com', 'david.smith@uscellular.com', 'Miguel.Jones@uscellular.com']

for who in sendTo:
    sendMail(excel_file, message, subject, who)
    
SystemExit(0);
