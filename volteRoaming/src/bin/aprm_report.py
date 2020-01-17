#! /usr/bin/python

# Oracle Libraries
import cx_Oracle

# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from dateutil.relativedelta import relativedelta
from datetime import datetime

import sys

######## Argument Date SQL Setup

date = "20200105" #sys.argv[1]

day = date[6:8]
month = date[4:6]
monthName = (datetime.strptime(date, '%Y%m%d')).strftime("%B")
year = date[0:4]

bp_start_date_LTE = (datetime.strptime(date[0:6] + "01", '%Y%m%d') - relativedelta(months=1)).strftime('%Y%m%d')
bp_start_date_CDMA = (datetime.strptime(date[0:6] + "16", '%Y%m%d') - relativedelta(months=1)).strftime('%Y%m%d')
end_date_cdma = ""

title = ""

if day == '05':
    end_date_cdma = " and  t1.sys_creation_date  < to_date('" + year + month + '01' + "','YYYYMMDD') "
    title = """The New!! APRM CDMA Accrual and 4G Settlement Report for """ + monthName + """ 5th """ + year 
    mess = " Attached to this email is the APRM CDMA Accrual and 4G Settlement Report for " + monthName + " 5th  " + year + ".\n The report covers the Billing Period Start Date of " + bp_start_date_LTE + " for 4G and " + bp_start_date_CDMA + " for CDMA.\n\nDave";

elif day == '22':
     title = """The New!! APRM CDMA Settlement Report for  """ + monthName + """ 22nd """ + year 
     mess = " Attached to this email is the APRM CDMA Settlement Report for " + monthName + " 22nd " + year + ".\n The report covers the Billing Period Start Date of " + bp_start_date_CDMA + ".\n\n Dave";

else:
    print("Incorrect date given\nMust be either the 5th or 22nd of the month\n")
    SystemExit(99);

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)

sqlDictionary = {}

sqlDictionary["LTE"] = """
  SELECT TO_CHAR (t1.sys_creation_date, 'YYYY-MM-DD')"Creation Date",
         t1.nr_param_3_val                          "Company Code",
         rate_plan_cd,
         CASE
             WHEN rate_plan_cd LIKE '%VoLTE%' AND t1.prod_cat_id = 'OS'
             THEN
                 '5438002'
             WHEN rate_plan_cd NOT LIKE '%VoLTE%' AND t1.prod_cat_id = 'OS'
             THEN
                 '5438001'
             WHEN rate_plan_cd LIKE '%VoLTE%' AND t1.prod_cat_id = 'IS'
             THEN
                 '6008002'
             WHEN rate_plan_cd NOT LIKE '%VoLTE%' AND t1.prod_cat_id = 'IS'
             THEN
                 '6008001'
             WHEN t1.rate_plan_cd = 'RPINCGSMDATACD' AND t1.prod_cat_id = 'II'
             THEN
                 '6008001'
             WHEN t1.rate_plan_cd = 'RPINCGSMSMSCD' AND t1.prod_cat_id = 'II'
             THEN
                 '6002202'
             WHEN     t1.rate_plan_cd = 'RPINCGSMVOICETOTCD'
                  AND t1.prod_cat_id = 'II'
             THEN
                 '6002201'
             ELSE
                 '0000000'
         END
             AS ascase,
         DECODE (carrier_cd,
                'USAAT', 'ATT Mobility (USAAT)',
                'USABS', 'ATT Mobility (USABS)',
                'USACC', 'ATT Mobility (USACC)',
                'USACG', 'ATT Mobility (USACG)',
                'USAMF', 'ATT Mobility (USAMF)',
                'USAPB', 'ATT Mobility (USAPB)',
                'USAKY', 'Appalachian Wireless (USAKY)',
                'USACM', 'C-Spire (USACM)',
                'USA1E', 'Carolina West (USA1E)',
                'USAXC', 'Inland (USAXC)',
                'USAJV', 'James Valley (USAJV)',
                'USA6G', 'Nex-Tech Wireless (USA6G)',
                'USAPI', 'Pioneer Cellular (USAPI)',
                'USASG', 'Sprint (USASG)',
                'USASP', 'Sprint (USASP)',
                'USASU', 'Sprint (USASU)',
                'USATM', 'T-Mobile (USATM)',
                'USAW6', 'T-Mobile (USAW6)',
                'GBRJT', 'Jersey Telecom (GBRJT)',
                'USA34', 'Illinois Valley Cell (USA34)',
                'USAUW', 'United Wireless (USAUW)',
                'USAVZ', 'Verizon (USAVZ)',
                'AAZVF', 'Vodafone Malta (AAZVF)',
                'MLTTL', 'Vodafone Malta (MLTTL)',
                'NLDLT', 'Vodafone Netherlands (NLDLT)',
                carrier_cd)
             "Carrier Code",
         t1.prod_cat_id,
         CASE
             WHEN    t1.rate_plan_cd = 'RPINCGSMVOICETOTCD'
                  OR t1.rate_plan_cd = 'RPINCGSMSMSCD'
             THEN
                 SUM (t1.TOT_CHRG_PARAM_VAL)
             ELSE
                 SUM ( (t1.TOT_CHRG_PARAM_VAL / 1024) / 1024)
         END
             AS Total_usage,
         SUM (t1.tot_net_usage_chrg)                "Total Charges"
    FROM IC_ACCUMULATED_USAGE t1
   WHERE     (   t1.prod_cat_id = 'IS'
              OR t1.prod_cat_id = 'OS'
              OR t1.prod_cat_id = 'II')
         AND t1.BP_START_DATE = TO_DATE ('""" + bp_start_date_LTE + """', 'YYYYMMDD')
GROUP BY TO_CHAR (t1.sys_creation_date, 'YYYY-MM-DD'),
         t1.nr_param_3_val,
         carrier_cd,
         rate_plan_cd,
         t1.prod_cat_id
ORDER BY TO_CHAR (t1.sys_creation_date, 'YYYY-MM-DD')
"""

sqlDictionary["CDMA_In"] = """
SELECT TO_CHAR (sys_creation_date, 'YYYY-MM-DD')"Create Date",
         carrier_cd,
         t1.rate_plan_cd,
         CASE
             WHEN t1.rate_plan_cd = 'RPROINDATA' THEN '6008001'
             WHEN t1.rate_plan_cd = 'RPROINVOIC' THEN '6002201'
             WHEN t1.rate_plan_cd = 'RPROINSRCHG' THEN '6002201'
             ELSE '0000000'
         END
             AS glcode,
         t2.carrier_name,
         CASE
             WHEN t1.prod_cat_id = 'IN' AND t1.rate_plan_cd = 'RPROINDATA'
             THEN
                 'ID'
             ELSE
                 t1.prod_cat_id
         END
             AS prod_id,
         CASE
             WHEN t1.rate_plan_cd = 'RPROINDATA'
             THEN
                 ( (SUM (t1.tot_chrg_param_val) / 1024) / 1024)
             ELSE
                 SUM (t1.tot_chrg_param_val)
         END
             AS total_vol,
         SUM (t1.tot_net_usage_chrg)             "Total Charges"
    FROM IC_ACCUMULATED_USAGE t1,
         (  SELECT setlmnt_contract_cd, MAX (Sid_Commercial_Name) carrier_name
              FROM pc9_sid
          GROUP BY setlmnt_contract_cd
          ORDER BY setlmnt_contract_cd) t2
   WHERE     SUBSTR (t1.nr_param_2_val, 0, 3) = t2.setlmnt_contract_cd
         AND (t1.prod_cat_id = 'IN')
         AND t1.BP_START_DATE = TO_DATE ('""" + bp_start_date_CDMA + """', 'YYYYMMDD')
         AND (t1.future_3 = 'Data' OR t1.future_3 = 'Voice') """ + end_date_cdma + """ 
GROUP BY t1.sys_creation_date,
         carrier_cd,
         t1.rate_plan_cd,
         t2.carrier_name,
         t1.prod_cat_id
ORDER BY t1.sys_creation_date, carrier_cd
""" 

sqlDictionary["CDMA_Out"] = """
SELECT TO_CHAR (sys_creation_date, 'YYYY-MM-DD')"Create Date",
         carrier_cd,
         t1.rate_plan_cd,
         CASE
             WHEN t1.rate_plan_cd = 'RPOUROAIR' THEN '5430001'
             WHEN t1.rate_plan_cd = 'RPOUROTOLL' THEN '5410101'
             WHEN t1.rate_plan_cd = 'RPROUROTAX' THEN '4080401'
             ELSE '0000000'
         END
             AS glcode,
         t2.carrier_name,
         t1.prod_cat_id,
                 SUM (t1.tot_chrg_param_val),
         SUM (t1.tot_net_usage_chrg)             "Total Charges"
    FROM IC_ACCUMULATED_USAGE t1,
         (  SELECT setlmnt_contract_cd, MAX (Sid_Commercial_Name) carrier_name
              FROM pc9_sid
          GROUP BY setlmnt_contract_cd
          ORDER BY setlmnt_contract_cd) t2
   WHERE     SUBSTR (t1.nr_param_1_val, 0, 3) = t2.setlmnt_contract_cd
         AND (t1.prod_cat_id = 'RO')
         AND t1.BP_START_DATE = TO_DATE ('""" + bp_start_date_CDMA + """', 'YYYYMMDD')
         AND (t1.future_3 = 'Voice') and t1.rate_plan_cd in ('RPOUROAIR','RPOUROTOLL','RPROUROTAX')
         AND t1.rate_plan_cd != 'RPOUROTOT' """ + end_date_cdma + """ 
GROUP BY t1.sys_creation_date,
         carrier_cd,
         t1.rate_plan_cd,
         t2.carrier_name,
         t1.prod_cat_id
ORDER BY t1.sys_creation_date, carrier_cd
""" 


def printRow (row, font, sheet, max_row):
    
    for i in range(0, len(row)):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font 
        sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00' 
    return;

 
def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "david.balchen@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("david.balchen@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def dbConnect ():
    
    CONN_INFO = {
         'host': '10.176.199.19',  # info from tnsnames.ora
         'port': 1530,  # info from tnsnames.ora
         'user': 'md1dbal1',
         'psw': 'P0tat000#',
         'service': 'bodsprd_adhoc'  # info from tnsnames.ora
    }

    CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)

    tconn = cx_Oracle.connect(CONN_STR)

#    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_SVC_BILLINGOPS")

    return tconn;


def returnSums (incoming, gl_codes):
    
    sums = []
    
    try:
        sums.append(sum([x[6] for x in incoming]) + 0.00)
        sums.append(sum([x[7] for x in incoming]) + 0.00)
    
        for glcode in gl_codes:
            sums.append(sum([x[6] for x in incoming if x[3] == glcode]) + 0.00)
            sums.append(sum([x[7] for x in incoming if x[3] == glcode]) + 0.00)
    except: pass   
    return sums


def processTableDate (res, gl_codes):
    
    out = []
    all_dates = sorted(set(map(lambda x:x[0], res)))
    
    
    
    for date in all_dates:
    
        company_dates = [x for x in res if x[0] == date ]  
    
        all_carrier = sorted(set(map(lambda x:x[1], company_dates))) 
    
        for carrier in all_carrier:
        
            incoming = [x for x in company_dates if x[1] == carrier]
            
            sums = returnSums(incoming, gl_codes)    
            
            if(incoming[0][5] == 'IS' or incoming[0][5] == 'OS'):   
                out.append((incoming[0][0], incoming[0][1], sums[0], sums[1], gl_codes[0], sums[3], gl_codes[1], sums[5]))
            elif(incoming[0][5] == 'II'):
                 out.append((incoming[0][0], incoming[0][1], "Vodofone Netherland", gl_codes[0], sums[2], sums[3], gl_codes[1], sums[4], sums[5], gl_codes[2], sums[6], sums[7]))  
            elif(incoming[0][5] == 'ID' or incoming[0][5] == 'IN'):
                out.append((incoming[0][0], incoming[0][1], gl_codes[0], sums[2], sums[3]))  
            elif(incoming[0][5] == 'RO'): 
                out.append((incoming[0][0], incoming[0][1], gl_codes[2], sums[2], sums[7], gl_codes[1], sums[5], gl_codes[0], sums[3]))     
              
    return out


def processTableCarrier (res, gl_codes):
    
    out = []
    company_codes = sorted(set(map(lambda x:x[1], res)))
    
    for company_code in company_codes:
        print("company code" + company_code + "\n")
        all_carriers = [x for x in res if x[1] == company_code ]
             
        carrier_by_company = sorted(set(map(lambda x:x[4], all_carriers)))
          
        for carrier in carrier_by_company:
            print("carrier " + carrier + "\n")       
            incoming = [x for x in all_carriers if x[4] == carrier]
                 
            sums = returnSums(incoming, gl_codes)
            
            if(incoming[0][5] == 'IS' or incoming[0][5] == 'OS'):                        
                out.append((incoming[0][1], carrier, sums[2], sums[3], sums[4], sums[5]))  
            elif(incoming[0][5] == 'ID' or incoming[0][5] == 'IN'):
              out.append((incoming[0][1], carrier, sums[2], sums[3])) 
            elif(incoming[0][5] == 'RO'):
              out.append((incoming[0][1], carrier, sums[6], gl_codes[2], sums[7], gl_codes[1], sums[5], gl_codes[0], sums[3]))     
              
    return out


def processTable (res, filterby='IS', processType='sum'):
    
    ires = [x for x in res if x[5] == filterby]
    
    gl_codes = sorted(set(map(lambda x:x[3], ires)))
    
#    gl_codes = [ x for x in gl_codes if "0000000" not in x ]
    
    if processType == 'sum':
        out = processTableDate (ires, gl_codes)
    
    elif processType == 'carrier':
        out = processTableCarrier (ires, gl_codes)
        
    else:
        print ("You Broke it!!!!")
        SystemExit(99);
    return out


def printSheet (title, header, sheet, output):
    sheet.title = title
    
    printRow(header, bold_font, sheet, 1)
    max_row = 2 

    for row in output:
        printRow(row, def_font, sheet, max_row)
        max_row = max_row + 1 
    return

###### Main Program  


excel_file = title + '.xlsx'
wb = Workbook()

conn = dbConnect()
cursor = conn.cursor()

results = []

if day == '05':
    
    print("Retrieve LTE data from Database\n")
    cursor.execute(sqlDictionary["LTE"])

    results = cursor.fetchall()
    
    print("Processing LTE Incollect Settlement\n")
    output = processTable(results, 'IS', 'sum')
    printSheet('LTE Incollect Settlement', ['Creation Date', 'Carrier Code', 'Total Usage MB', 'Total Charges', 'GL Data', 'Data Charges', 'GL VoLTE', 'VoLTE Charges'], wb.active, output)
    
    print("Processing LTE Incollect by Carrier\n")
    output = processTable(results, 'IS', 'carrier')
    printSheet('LTE Incollect by Carrier', ['Carrier Code', 'Carrier', 'Total LTE Usage MB', 'Total LTE Charges', 'Total VoLTE Usage MB', 'Total VoLTE Charges'], wb.create_sheet("LTE Incollect by Carrier"), output)
    
    print("Processing LTE Outcollect Settlement\n")
    output = processTable(results, 'OS', 'sum')
    printSheet('LTE Outcollect Settlement', ['Creation Date', 'Carrier Code', 'Total Usage MB', 'Total Charges', 'GL Data', 'Data Charges', 'GL VoLTE', 'VoLTE Charges'], wb.create_sheet("LTE Outcollect Settlement"), output)
    
    print("Processing LTE Outcollect by Carrier\n") 
    output = processTable(results, 'OS', 'carrier')
    printSheet('LTE Outcollect by Carrier', ['Carrier Code', 'Carrier', 'Total LTE Usage MB', 'Total LTE Charges', 'Total VoLTE Usage MB', 'Total VoLTE Charges'], wb.create_sheet("LTE Outcollect by Carrier"), output)
    
    print("Processing GSM Incollect Settlement\n")  
    output = processTable(results, 'II', 'sum')
    printSheet('GSM Incollect Settlement', ['Creation Date', 'Company Code', 'Carrier', 'VOICE GL', 'Total Minutes', 'Voice Charges', 'TEXT GL', 'Total Texts', 'Text Charges', 'DATA GL', 'Total Usage MB', 'Data Charges'], wb.create_sheet("GSM Incollect Settlement"), output)
     
    print("Retrieve CDMA Incollect data from Database\n")
    cursor.execute(sqlDictionary["CDMA_In"])
    
    results = cursor.fetchall()
    
    print("Processing CDMA Data Incollect Accrual\n")   
    output = processTable(results, 'ID', 'sum')
    printSheet('CDMA Data Incollect Accrual', ['Creation Date', 'Carrier Code', 'GL Account', 'Total Usage MB', 'Total Charges'], wb.create_sheet("CDMA Data Incollect Accrual"), output)
    
    print("Processing CDMA Data Incollect by Carrier\n") 
    output = processTable(results, 'ID', 'carrier')
    printSheet('CDMA Data Incollect by Carrier', ['Carrier Code', 'Carrier Name', 'Total Usage MB', 'Total Charges'], wb.create_sheet("CDMA Data Incollect by Carrier"), output)
    
    print("Processing CDMA Voice Incollect Accrual\n") 
    output = processTable(results, 'IN', 'sum')
    printSheet('CDMA Voice Incollect Accrual', ['Creation Date', 'Carrier Code', 'GL Account', 'Total Usage Minutes', 'Total Charges'], wb.create_sheet("CDMA Voice Incollect Accrual"), output)
   
    print("Processing CDMA Voice Incollect by Carrier\n")   
    output = processTable(results, 'IN', 'carrier')
    printSheet('CDMA Voice Incollect by Carrier', ['Carrier Code', 'Carrier Name', 'Total Usage Minutes', 'Total Charges'], wb.create_sheet("CDMA Voice Incollect by Carrier"), output)
    
    print("Retrieve CDMA Outcollect data from Database\n")
    cursor.execute(sqlDictionary["CDMA_Out"])
    results = cursor.fetchall()
    
    print("Processing CDMA Voice Outcollect Accrual\n")
    output = processTable(results, 'RO', 'sum')
    printSheet('CDMA Voice Outcollect Accrual', ['Creation Date', 'Carrier Code', 'GL Account', 'Total Usage Minutes', 'Total Charges Air', 'GL Account Toll', 'Total Charges Toll', 'GL Account Tax', 'Total Charges Tax'], wb.create_sheet("CDMA Voice Outcollect Accrual"), output)
 
    print("Processing CDMA Voice Outcollect Carrier\n")    
    output = processTable(results, 'RO', 'carrier')
    printSheet('CDMA Voice Outcollect Carrier', ['Carrier Code', 'Carrier Name', 'Total Usage Minutes', 'Air GL Code', 'Total Air Charges', 'Toll GL Code', 'Total Toll Charges', 'Tax GL Code', 'Total Tax Charges'], wb.create_sheet("CDMA Voice Outcollect Carrier"), output)
    
elif day == '22':
    
    print("Retrieve CDMA Incollect data from Database\n")    
    cursor.execute(sqlDictionary["CDMA_In"])
    
    print(sqlDictionary["CDMA_In"])
    
    results = cursor.fetchall()
    
    print("Processing CDMA Data Incollect Settlement\n")   
    output = processTable(results, 'ID', 'sum')
    printSheet('CDMA Data Incollect Settlement', ['Creation Date', 'Carrier Code', 'GL Account', 'Total Usage MB', 'Total Charges'], wb.active, output)
     
    print("Processing CDMA Data Incollect by Carrier\n")     
    output = processTable(results, 'ID', 'carrier')
    printSheet('CDMA Data Incollect by Carrier', ['Carrier Code', 'Carrier Name', 'Total Usage MB', 'Total Charges'], wb.create_sheet("CDMA Data Incollect by Carrier"), output)

    print("Processing CDMA Voice Incollect Settlement\n")   
    output = processTable(results, 'IN', 'sum')
    printSheet('CDMA Voice Incollect Settlement', ['Creation Date', 'Carrier Code', 'GL Account', 'Total Usage Minutes', 'Total Charges'], wb.create_sheet("CDMA Voice Incollect Settlement"), output)
 
    print("Processing CDMA Voice Incollect by Carrier\n")    
    output = processTable(results, 'IN', 'carrier')
    printSheet('CDMA Voice Incollect by Carrier', ['Carrier Code', 'Carrier Name', 'Total Usage Minutes', 'Total Charges'], wb.create_sheet("CDMA Voice Incollect by Carrier"), output)
 
    print("Retrieve CDMA Outcollect data from Database\n") 
    cursor.execute(sqlDictionary["CDMA_Out"])
    results = cursor.fetchall()
 
    print("Processing CDMA Voice Outcollect Settle\n") 
    output = processTable(results, 'RO', 'sum')
    printSheet('CDMA Voice Outcollect Settle', ['Creation Date', 'Carrier Code', 'GL Account', 'Total Usage Minutes', 'Total Charges Air', 'GL Account Toll', 'Total Charges Toll', 'GL Account Tax', 'Total Charges Tax'], wb.create_sheet("CDMA Voice Outcollect Settlem"), output)
 
    print("Processing CDMA Voice Outcollect Carrier\n")   
    output = processTable(results, 'RO', 'carrier')
    printSheet('CDMA Voice Outcollect Carrier', ['Carrier Code', 'Carrier Name', 'Total Usage Minutes', 'Air GL Code', 'Total Air Charges', 'Toll GL Code', 'Total Toll Charges', 'Tax GL Code', 'Total Tax Charges'], wb.create_sheet("CDMA Voice Outcollect Carrier"), output)
     
cursor.close
conn.close()

wb.save(excel_file)

message = mess
subject = title
#sendTo = ["david.balchen@uscellular.com",    'ISBillingOperations@uscellular.com','Ilham.Elgarni@uscellular.com','Heather.Jeschke@uscellular.com','david.smith@uscellular.com','Miguel.Jones@uscellular.com']

sendTo = ["david.balchen@uscellular.com"]

for who in sendTo:
    sendMail(excel_file, message, subject, who)
    
SystemExit(0);


