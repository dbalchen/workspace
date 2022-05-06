#! /usr/bin/python3

'''
Created on Jan 11, 2021
@author: dbalchen

'''
# Oracle Libraries
import cx_Oracle

import fileinput
import sys
# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from dateutil.relativedelta import relativedelta
from datetime import datetime

sendTo = ["david.balchen@uscellular.com", "Marvin.Guss@uscellular.com", "RevenueAccounting@uscellular.com", "xavier.lbataille@uscellular.com", "mark.foster@uscellular.com", "gabe.hedstrom@uscellular.com", "dean.schempp@uscellular.com", "Sandra.Fitts@uscellular.com", "olivia.solis@uscellular.com"]
#sendTo = ["david.balchen@uscellular.com"]

sqlDictionary = {}

sqlDictionary["OnNet"] = """
select to_char(a.session_start_time,'YYYY-MM-DD') "Session Date",    
    DECODE (switch_id,
    'MADI','Madison',
    'YAKI', 'Yakima',
    'KNOX', 'Knoxville',
    'PEO2', 'Peoria',
    'GREE', 'Greenville',
    'CDR2', 'Cedar Rapids',
    'OKLA', 'Oklahoma City',
    'NEWB', 'New Berlin',
    'MORG', 'Morgantown',
    'JOPL', 'Joplin',
    'OWAS', 'Owasso',
    'CONG', 'Congress Park',
    'ROC2', 'Rockford',
    'ASHE', 'Asheville',
    'MEDF', 'Medford',
    'LLYN', 'Lynchburg',
    'GRAN', 'Granville',
    'SALI', 'Salina',
    'CLIN', 'Clinton',
    'COLU', 'Columbia',
    'LROE', 'Roanoke',
    'EURE', 'Eureka',
    'STLO', 'St. Louis',
    'APPL', 'Appleton',
    'OMAH', 'Omaha',
    'JOHN', 'Johnstown',
    switch_id) "Switch ID",
    {sitename}
    DECODE (switch_id,
    'MADI','WI',
    'YAKI','WA',
    'KNOX', 'TN',
    'PEO2', 'IL',
    'GREE', 'NC',
    'CDR2', 'IA',
    'OKLA', 'OK',
    'NEWB', 'WI',
    'MORG', 'KY',
    'JOPL', 'MO',
    'OWAS', 'OK',
    'CONG', 'NH',
    'ROC2', 'IL',
    'ASHE', 'NC',
    'MEDF', 'OR',
    'LLYN', 'VA',
    'GRAN', 'ME',
    'SALI', 'KS',
    'CLIN', 'NC',
    'COLU', 'MO',
    'LROE', 'VA',
    'EURE', 'CA',
    'STLO', 'MO',
    'APPL', 'WI',
    'OMAH', 'NB',
    'JOHN', 'IA',
    switch_id) "Switch State",
    count(*) "Total Records", sum(a.session_duration) "Total Duration", sum(a.bytes_sent + a.bytes_received)/1024/1024 "Total MB" 
    from CDMA_AAA_DATA_USAGE a, BSID_TO_SERVE_SID@BRMPRD b
    where substr(a.bsid,0,11) = b.bsid 
    and a.session_start_time >=  to_date('{timeStamp}','YYYYMMDD') and a.session_start_time <  to_date('{endTimeStamp}','YYYYMMDD')
    group by to_char(a.session_start_time,'YYYY-MM-DD'), b.switch_id {sitename2}
    order by to_char(a.session_start_time,'YYYY-MM-DD'), b.switch_id {sitename2}
 """
 
sqlDictionary["ciber32"] = """
 select
  '20' || call_date "Call Date",
  count(*) "Record Count",
  b.sid_commercial_name "Carrier Name",
  b.sid_city "Service Place",
  b.sid_state "Service State",
  b.sid_country "Service Country",
  0.00 "MB Sent",
  0.00 "MB Received" ,
  0.00 "Total Real Usage",
  '6008001' "GL Account",
  to_char(
     sum(message_accounting_digits)/ 1024 / 1024,
     'fm9999990.90'
     ) "Total Charged Usage MB by State/Place",
  to_char(
     sum(total_charges_and_taxes)/ 100,
     'fm9999990.90'
     ) "Total Charges and Taxes by State/Place ",
  to_char(
     sum(
       total_state_taxes + total_local_taxes
         )/ 100,
     'fm9999990.90'
     ) "Total Taxes by State/Place"
from
  CDMA_CIBER32_DATA_USAGE a,
  pc9_sid b
where
 (to_date(call_date,'YYMMDD') >= to_date('{timeStamp}','YYYYMMDD')
 and  to_date(call_date,'YYMMDD') <  to_date('{endTimeStamp}','YYYYMMDD'))
 and lpad(
  a.serving_carrier_sid_bid, 5, '0'
  ) = b.sids and
  (b.expiration_date >= to_date(call_date,'YYMMDD'))
group by
  call_date,
  b.sid_commercial_name,
  b.sid_city,
  b.sid_state,
  b.sid_country
order by
  call_date,
  b.sid_commercial_name,
  b.sid_city,
  b.sid_state,
  b.sid_country
"""


sqlDictionary["usagebycarrier"] = """
select to_char(c.session_start_time,'YYMMDD') "Call Date",
  d.sid_commercial_name "Carrier Name",
  d.sid_city "Service Place",
  d.sid_state "Service State",
  d.sid_country "Service Country",
  to_char(
  sum(c.bytes_sent) /1024 / 1024 ,
  'fm9999990.9000')
  "MB Sent",
  to_char(
  sum(c.bytes_received) /1024 /1024,  
    'fm9999990.9000')
  "MB Received",
  to_char(
  (sum(c.bytes_sent) + sum(c.bytes_received)) /1024 /1024 ,
     'fm9999990.9000') "MB Total"
from CDMA_AAA_DATA_USAGE c, 
pc9_sid d
where session_start_time >= to_date('{timeStamp}','YYYYMMDD') and session_start_time < to_date('{endTimeStamp}','YYYYMMDD')
and lpad(
  to_number(
    substr(c.bsid,1,4),'XXXXX'
    )
    , 5,'0'
    ) = d.sids
and (d.expiration_date >= c.session_start_time)
and d.sid_commercial_name != 'U.S. Cellular'
group by to_char(c.session_start_time,'YYMMDD'),
  d.sid_commercial_name,
  d.sid_city,
  d.sid_state,
  d.sid_country
order by to_char(c.session_start_time,'YYMMDD'),
  d.sid_commercial_name,
  d.sid_city,
  d.sid_state,
  d.sid_country
"""


sqlDictionary["ciber22"] = """
select 
  '20' || call_date "Call Date", 
  count(*) "Record Count", 
  b.sid_commercial_name "Carrier Name", 
  b.sid_city "Service Place", 
  b.sid_state "Service State", 
  serving_country "Service Country", 
  '6002201' "GL Account", 
  to_char(
    sum(air_elapsed_time / 60), 
    'fm9999990.90'
  ) "Total Usage Minutes By State/Place", 
  to_char(
    sum(total_charges_and_taxes)/ 100, 
    'fm9999990.90'
  ) "Total Charges and Taxes by State/Place ", 
  to_char(
    sum(
      total_state_taxes + total_local_taxes
    )/ 100, 
    'fm9999990.90'
  ) "Total Taxes by State/Place" 
from 
  CDMA_CIBER22_VOICE_USAGE a, 
  pc9_sid b  
where  (to_date(call_date,'YYMMDD') >= to_date('{timeStamp}','YYYYMMDD') 
 and  to_date(call_date,'YYMMDD') <  to_date('{endTimeStamp}','YYYYMMDD'))
  and 
  lpad(
    a.serving_carrier_sid_bid, 5, '0'
  ) = b.sids 
group by 
  call_date, 
  b.sid_commercial_name, 
  b.sid_city, 
  b.sid_state, 
  serving_country 
order by 
  call_date, 
  b.sid_commercial_name, 
  b.sid_city, 
  b.sid_state, 
  serving_country
"""

headings = {}
 
headings["OnNet"] = [
"Session Date",
"Switch ID",
"Switch State",
"Total Records",
"Total Duration",
"Total MB"
];

headings["ciber32"] = [
"Call Date",
"Record Count",
"Carrier Name",
"Service Place",
"Service State",
"Service Country",
"MB In",
"MB Out",
"Total Real MB",
"GL Account",
"Total Usage MB by State/Place",
"Total Charges and Taxes by State/Place",
"Total Taxes by State/Place"
];

headings["ciber22"] = [
"Call Date",
"Record Count",
"Carrier Name",
"Service Place",
"Service State",
"Service Country",
"GL Account",
"Total Usage Minutes By State/Place",
"Total Charges and Taxes by State/Place",
"Total Taxes by State/Place"
];

tab = {}
 
tab["OnNet"] = "OnNet Data Usage";
tab["ciber22"] = "Off-net Voice (Ciber 22)";
tab["ciber32"] = "Off-net Data (Ciber 32)";

 
def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "david.balchen@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("david.balchen@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def dbConnect ():
    
    CONN_INFO = {
        'host': '10.176.199.19',  # info from tnsnames.ora
        'port': 1530,  # info from tnsnames.ora
        'user': 'md1dbal1',
        'psw': 'XXXXXXXXXXXXXXXX',
        'service': 'bodsprd_adhoc'  # info from tnsnames.ora
    }
    CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)
   
#    tconn = cx_Oracle.connect(CONN_STR)

    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_SVC_BILLINGOPS")
    
    return tconn;


def printRow (row, font, sheet, max_row, max_col=0): 
    for i in range(0, len(row) - max_col):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font
        if i > 2: 
            sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00'             
    return;


def printSheet (title, header, sheet, output, flag=0):
    sheet.title = title
    
    printRow(header, bold_font, sheet, 1)
    max_row = 2 

    for row in output:
        font = def_font

        # Modify fonts here

        printRow(row, font, sheet, max_row, (len(row) - len(header)))
        max_row = max_row + 1 
    return

if __name__ == '__main__':
    pass

timeStamp = '20220401' # sys.argv[1]

endTimeStamp = timeStamp[0:6] + "01"
timeStamp = endTimeStamp

thisMonth = (datetime.strptime(endTimeStamp, '%Y%m%d')).strftime("%B")

timeStamp = datetime.strptime(timeStamp, "%Y%m%d")
timeStamp = (timeStamp - relativedelta(months=1)).strftime('%Y%m%d')

day = "1"
month = timeStamp[4:6]
monthName = (datetime.strptime(timeStamp, '%Y%m%d')).strftime("%B")
year = timeStamp[0:4]

sitename = ""
sitename2 = ""

title = """The Aeris Report for """ + thisMonth + """ """ + day + """, """ + year 
message = " Attached to this email is the Aeris Report for " + thisMonth + """ """ + day + """, """ + year + "\n\n"

# Open an Excel file for output

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=False)

excel_file = title + '.xlsx'
wb = Workbook()

# Database

# conn = dbConnect()
# cursor = conn.cursor()
 
# Do CIBER22
results = []
sql = sqlDictionary["ciber22"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp, sitename=sitename, sitename2=sitename2)
print(sql)
 
# cursor.execute(sql)
# results = cursor.fetchall()
#  

printSheet(tab["ciber22"], headings["ciber22"], wb.active, results, 1);

# Do Ciber32

results = []
sql = sqlDictionary["ciber32"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp, sitename=sitename, sitename2=sitename2)
print(sql)

# cursor.execute(sql)
# results = cursor.fetchall()

# for line in fileinput.input("/home/dbalchen/Desktop/test.csv"):
#     try:
#         line = line.rstrip()
#         results.append(tuple(line.split("\t")))
#     except:pass

results2 = []
sql = sqlDictionary["usagebycarrier"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp, sitename=sitename, sitename2=sitename2)
print(sql)

# cursor.execute(sql)
# results2 = cursor.fetchall()

# for line in fileinput.input("/home/dbalchen/Desktop/test2.csv"):
#     try:
#         line = line.rstrip()
#         results2.append(tuple(line.split("\t")))
#     except:pass

# for idx,crec in enumerate(results):
#     
#     crec = list(crec)
#     
#     date = (crec[0])[2:8] 
#     rec_dates = [x for x in results2 if x[0] == date] 
#     
#     cc = crec[2]
#     rec_carrier = [x for x in rec_dates if x[1] == cc] 
#     
#     ccty = crec[3]
#     carrier_city = [x for x in rec_carrier if x[2] == ccty] 
#     
#     cst = crec[4]
#     carrier_state = [x for x in carrier_city if x[3] == cst] 
#     
#     cco = crec[5]
#     carrier_country = [x for x in carrier_state if x[4] == cco]
#     
#     try :
#         crec[6] = (carrier_country[0])[5]
#         crec[7] = (carrier_country[0])[6]
#         crec[8] = (carrier_country[0])[7]
#         
#         results[idx] = tuple(crec)
#     except:
#         pass
# 
# 
# 
# printSheet(tab["ciber32"], headings["ciber32"], wb.create_sheet(tab["ciber32"]), results, 1);


# Do AAA Processing378004
results = []
sql = sqlDictionary["OnNet"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp, sitename=sitename, sitename2=sitename2)
# 
print(sql)
#     
# cursor.execute(sql)
# results = cursor.fetchall()
# 
printSheet(tab["OnNet"], headings["OnNet"], wb.create_sheet(tab["OnNet"]), results, 1);

wb.save(excel_file)

# Close database connection
cursor.close()
conn.close()


for who in sendTo:
    sendMail(excel_file, message, title, who)
    
SystemExit(0);

