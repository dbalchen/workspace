#! /usr/bin/python3

'''
Created on April 28, 2022
@author: dbalchen

'''
# Oracle Libraries
import cx_Oracle

import fileinput
import sys
# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from dateutil.relativedelta import relativedelta
from datetime import datetime

#sendTo = ["david.balchen@uscellular.com", "Marvin.Guss@uscellular.com", "michael.joseph@uscellular.com", "xavier.lbataille@uscellular.com", "mark.foster@uscellular.com", "gabe.hedstrom@uscellular.com", "dean.schempp@uscellular.com", "Sandra.Fitts@uscellular.com", "olivia.solis@uscellular.com"]
sendTo = ["david.balchen@uscellular.com"]

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=False)

sqlDictionary = {}


sqlDictionary["IOT_RATE_PLAN"] = """
  SELECT *
  FROM IOT_RATE_PLAN t1,
       (  SELECT t2.start_date, MIN (ABS (t2.start_date - to_date('{this_month}','YYYYMMDD'))) diff
            FROM IOT_RATE_PLAN t2
        GROUP BY t2.start_date) b
 WHERE ABS (t1.start_date - to_date('{this_month}','YYYYMMDD')) = (b.diff)
 and t1.start_date <= to_date('{this_month}','YYYYMMDD')
 and (t1.end_date >= to_date('{this_month}','YYYYMMDD') 
      or t1.end_date is NULL)
 order by TADIG, IMSI_BILLING_TYPE, PLAN_ID, t1.start_date desc
"""

sqlDictionary["IOT_PARTNER"] =  """
select unique(tadig),mvno_name from IOT_PARTNER
"""

# sqlDictionary["IOT_IMSI_STATUS"] = """
#     select IMSI, TADIG, ACTIVATED_DATE, DEACTIVATED_DATE  from IOT_IMSI_STATUS where ACTIVATED_DATE  >= to_date('{last_month}','YYMMDD') and (DEACTIVATED_DATE <= to_date('{this_month}','YYMMDD') or DEACTIVATED_DATE = '00:00:00')
# """

sqlDictionary["IOT_IMSI_RANGE"] = """
select * from IOT_IMSI_RANGE t1,
 (  SELECT t2.start_date, MIN (ABS (t2.start_date - to_date('{this_month}','YYYYMMDD'))) diff
            FROM IOT_IMSI_RANGE t2
        GROUP BY t2.start_date) b
 WHERE ABS (t1.start_date - to_date('{this_month}','YYYYMMDD')) = (b.diff)
 and t1.start_date <= to_date('{this_month}','YYYYMMDD')
 and (t1.end_date >= to_date('{this_month}','YYYYMMDD') 
      or t1.end_date is NULL) 
order by t1.TADIG, t1.IMSI_BILLING_TYPE, t1.start_DATE desc
"""

sqlDictionary["usage"] = """
select IMSI,TADIG,count(*), sum(TOTAL_CALL_EVENT_DURATION)/60, (sum(DATA_VOLUME_INCOMING)/1024),
       sum(DATA_VOLUME_OUTGOING/1024),sum((DATA_VOLUME_INCOMING/1024) + DATA_VOLUME_OUTGOING/1024),sum(CHARGE) 
 from  {IOT_AGGREGATOR}
 where CALL_EVENT_START_TIMESTAMP >= to_date('{last_month}','YYYYMMDD') and CALL_EVENT_START_TIMESTAMP < to_date('{this_month}','YYMMDD')
 group by IMSI,TADIG  
"""

headings = {}

headings["IoT_IMSI"] = [
"IMSI",
"Total Records",
"Total Session Duration (Minutes)",
"Incoming KB",
"Outgoing KB",
"Total Record Charges $",
"Plan A Charged Units KB",
"Plan B Charged Units KB"
];

tab = {}
tab["IoT_IMSI"] = "IoT {tadig} by IMSI {onOff}"; 



def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "david.balchen@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("david.balchen@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def dbConnect ():
    
    CONN_INFO = {
        'host': '10.180.171.15',  # info from tnsnames.ora
        'port': 1530,  # info from tnsnames.ora
        'user': 'md1dbal1',
        'psw': 'Potat000#',
        'service': 'bodsdev_adhoc'  # info from tnsnames.ora
    }
    CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)
  
#    tconn = cx_Oracle.connect(CONN_STR)

    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_SVC_BILLINGOPS")
    
    return tconn;


def printRow (row, font, sheet, max_row, max_col=0): 
    for i in range(0, len(row) - max_col):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font
        if i > 2: 
            sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00'             
    return;


def printSheet (title, header, sheet, output, flag=0):
    sheet.title = title
    
    printRow(header, bold_font, sheet, 1)
    max_row = 2 

    for row in output:
        font = def_font

        # Modify fonts here

        printRow(row, font, sheet, max_row, (len(row) - len(header)))
        max_row = max_row + 1 
    return


######################################################################################
# 

def  processUsage(wb,usage,onOff = 'OnNet',count = 0):
    
    tadig = sorted(set(map(lambda x:x[1], usage)))
    
    for tad in tadig: 
       
        mvno = [x for x in iotPartner if x[0] == tad][0][1]
    
        results = [x for x in usage if x[1] == tad]
        
        output = []
        
        for row in results:
        
            imsi = row[0]
        
            try :
                planID = [x for x in iot_imsi_range if (x[0] == tad and (imsi >= x[2] and imsi <= x[3])) ][0][1]
            except :
                print("IMSI not in  line range for tadig\n",file=sys.stderr)
                print(row,file=sys.stderr)  
                print("\n",file=sys.stderr)    
                continue   
        
            try :                    
                rate_plan = [x for x in iot_rate_plan if (x[0] == tad and x[1] == planID)][0]
            except :
                    print("No Rate Plan for tadig\n  :",file=sys.stderr)
                    print(row,file=sys.stderr) 
                    print("\n",file=sys.stderr)   
                    continue        
             
            if onOff == 'OnNet' : 
                totNetCharges = float(rate_plan[6])*float(row[6]);
            else :
                totNetCharges = float(rate_plan[5])*float(row[6]);
                       
            if row[6] <= rate_plan[3] :             
                rowTuple = (imsi,int(row[2]),float(row[3]),float(row[4]),float(row[5]),float(totNetCharges),float(row[6]),'')          
            else:  
                rowTuple = (imsi,int(row[2]),float(row[3]),float(row[4]),float(row[5]),float(totNetCharges),'',float(row[6]))             
            
            output.append(rowTuple)
    
        if count == 0 :
            printSheet(tab["IoT_IMSI"].format(tadig=tad,onOff=onOff ), headings["IoT_IMSI"], wb.active, output, 1);
        else:         
            printSheet(tab["IoT_IMSI"].format(tadig=tad,onOff=onOff ), headings["IoT_IMSI"], wb.create_sheet(tab["IoT_IMSI"].format(tadig=tad,onOff=onOff )), output, 1);
        
        count = count + 1


    return count

if __name__ == '__main__':
    pass


# Date setup

timeStamp = "20220501" # sys.argv[1]
this_month = timeStamp[0:6] + "01"
year = this_month[0:4]
monthName = (datetime.strptime(this_month, '%Y%m%d')).strftime("%B")

last_month = datetime.strptime(this_month, "%Y%m%d")
last_month = (last_month - relativedelta(months=1)).strftime('%Y%m%d')

# Email title and message
title = """The IoT Usage Report for """ + monthName + """ 1st """ + """, """ + year 
message = "Attached to this email is the IoT Usage Report for " + monthName + """ 1st """ + """, """ + year + "\n\n"

# Open an Excel file for output

excel_file = title + '.xlsx'
wb = Workbook()

# Database

# conn = dbConnect()
# cursor = conn.cursor()
#  

# Load Reference tables into a list.

iotPartner = []
sql = sqlDictionary["IOT_PARTNER"].format(last_month=last_month, this_month=this_month)
print(sql)
 
# cursor.execute(sql)
# iotPartner = cursor.fetchall()
 
for line in fileinput.input("/home/dbalchen/Desktop/IOT_PARTNER.csv"):
    try:
        line = line.rstrip()
        iotPartner.append(tuple(line.split("\t")))
    except:pass


iot_imsi_range = []
sql = sqlDictionary["IOT_IMSI_RANGE"].format(last_month=last_month, this_month=this_month)
print(sql)
 
# cursor.execute(sql)
# iot_imsi = cursor.fetchall()

for line in fileinput.input("/home/dbalchen/Desktop/IOT_IMSI_RANGE.csv"):
    try:
        line = line.rstrip()
        iot_imsi_range.append(tuple(line.split("\t")))
    except:pass


iot_rate_plan = []
sql = sqlDictionary["IOT_RATE_PLAN"].format(last_month=last_month, this_month=this_month)
print(sql)
 
# cursor.execute(sql)
# iot_rate_plan = cursor.fetchall()

for line in fileinput.input("/home/dbalchen/Desktop/IOT_RATE_PLAN.csv"):
    try:
        line = line.rstrip()
        iot_rate_plan.append(tuple(line.split("\t")))
    except:pass


usage = []
sql = sqlDictionary["usage"].format(last_month=last_month, this_month=this_month,IOT_AGGREGATOR = 'IOT_AGGREGATOR_ON_NET_USAGE')
print(sql)

onOff = 'OnNet' # example
 
# cursor.execute(sql)
# usage = cursor.fetchall()

for line in fileinput.input("/home/dbalchen/Desktop/IOT_AGGREGATOR_ON_NET_USAGE.csv"):
    try:
        line = line.rstrip()
        usage.append(tuple(line.split("\t")))
    except:pass


count = processUsage(wb,usage,onOff)


usage = []
sql = sqlDictionary["usage"].format(last_month=last_month, this_month=this_month,IOT_AGGREGATOR = 'IOT_AGGREGATOR_OFF_NET_USAGE')
print(sql)

onOff = 'OffNet' # example
 
# cursor.execute(sql)
# usage = cursor.fetchall()

for line in fileinput.input("/home/dbalchen/Desktop/IOT_AGGREGATOR_ON_NET_USAGE.csv"):
    try:
        line = line.rstrip()
        usage.append(tuple(line.split("\t")))
    except:pass


count = processUsage(wb,usage,onOff,count)

wb.save(excel_file)

# Close database connection

# cursor.close()
 
# conn.close()

#    Send report to our business partners.
for who in sendTo:
    sendMail(excel_file, message, title, who)
    
SystemExit(0);

