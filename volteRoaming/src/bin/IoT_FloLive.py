#! /usr/bin/python3

'''
Created on Jan 11, 2021
@author: dbalchen

'''
# Oracle Libraries
import cx_Oracle

import fileinput
import sys
# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from dateutil.relativedelta import relativedelta
from datetime import datetime

#sendTo = ["david.balchen@uscellular.com", "Marvin.Guss@uscellular.com", "michael.joseph@uscellular.com", "xavier.lbataille@uscellular.com", "mark.foster@uscellular.com", "gabe.hedstrom@uscellular.com", "dean.schempp@uscellular.com", "Sandra.Fitts@uscellular.com", "olivia.solis@uscellular.com"]
#sendTo = ["david.balchen@uscellular.com"]
sendTo = ["david.balchen@uscellular.com", "Michael.Joseph@uscellular.com","USCDLRA-Monitoring&Metrics@uscellular.com","Christine.Bekos@uscellular.com","Marvin.Guss@uscellular.com", "RevenueAccounting@uscellular.com", "xavier.lbataille@uscellular.com", "mark.foster@uscellular.com", "gabe.hedstrom@uscellular.com", "dean.schempp@uscellular.com", "Sandra.Fitts@uscellular.com", "olivia.solis@uscellular.com"]

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=False)

sqlDictionary = {}


sqlDictionary["IoT"] = """
select imsi "IMSI",
count(*) "Total Records",
serving_location_description "Serving Location",
Decode (sender_id, 'USAAT', 'ATT Mobility (USAAT)',
'USABS', 'ATT Mobility (USABS)',
'USACC', 'ATT Mobility (USACC)',
'USACG', 'ATT Mobility (USACG)',
'USAMF', 'ATT Mobility (USAMF)',
'USAPB', 'ATT Mobility (USAPB)',
'USAKY', 'Appalachian Wireless (USAKY)',
'USACM', 'C-Spire (USACM)',
'USA1E', 'Carolina West (USA1E)',
'USAXC', 'Inland (USAXC)',
'USAJV', 'James Valley (USAJV)',
'USA6G', 'Nex-Tech Wireless (USA6G)',
'USAPI', 'Pioneer Cellular (USAPI)',
'USASG', 'Sprint (USASG)',
'USASP', 'Sprint (USASP)',
'USASU', 'Sprint (USASU)',
'USATM', 'T-Mobile (USATM)',
'USAW6', 'T-Mobile (USAW6)',
'GBRJT', 'Jersey Telecom (GBRJT)',
'USA34', 'Illinois Valley Cell (USA34)',
'USAUW', 'United Wireless (USAUW)',
'USAVZ', 'Verizon (USAVZ)',
'AAZVF', 'Vodafone Malta (AAZVF)',
'MLTTL', 'Vodafone Malta (MLTTL)',
'NLDLT', 'Vodafone Netherlands (NLDLT)',
sender_ID)             "Roaming Partner",
to_char(
sum(Total_call_event_duration)/60,
'fm9999999990.90'
)
"Total Session Duration (Minutes)",
to_char(
  sum(data_volume_incoming)/1024 ,
    'fm9999999990.9000'
  ) 
"Incoming KB",
to_char(
   sum(data_volume_outgoing)/1024 ,
    'fm9999999990.9000'
  ) 
"Outgoing KB", 
to_char(
   (sum(data_volume_incoming)+ sum(data_volume_outgoing))/1024 ,
    'fm9999999990.9000'
  ) 
"Total Volume KB",
to_char(
   sum(charge) ,
    'fm9999999990.9000'
  ) 
"Total Record Charges $",
to_char(  
        sum(Charged_Units)/1024,'fm9999999990.90000'
  ) 
 "Charged Units KB"
from IOT_AGGREGATOR_ON_NET_USAGE where tadig = 'IOTFL'
and (ods_insert_date >= to_date('{timeStamp}','YYYYMMDD') 
and ods_insert_date < to_date('{endTimeStamp}','YYYYMMDD') )
group by
imsi,sender_id,serving_location_description
order by
imsi,sender_id, serving_location_description
"""


sqlDictionary["IoT_IMSI"] = """
select imsi "IMSI",
count(*) "Total Records",
to_char(
sum(Total_call_event_duration)/60,
'fm9999999990.90'
)
"Total Session Duration (Minutes)",
to_char(
   sum(data_volume_incoming)/1024 ,
    'fm9999999990.9000'
  ) 
"Incoming KB",
to_char(
   sum(data_volume_outgoing)/1024 ,
    'fm9999999990.9000'
  ) 
"Outgoing KB", 
to_char(
   (sum(data_volume_incoming)+ sum(data_volume_outgoing))/1024 ,
    'fm9999999990.9000'
  ) 
"Total Volume KB",
to_char(
   sum(charge) ,
    'fm9999999990.900000'
  ) 
"Total Record Charges $",
to_char(  
            CASE WHEN  sum(Charged_Units)/1024 >= 77824
            then 0.00
         Else sum(Charged_Units)/1024
         End  
   ,
    'fm9999999990.9000'
  ) 
 "Plan A Charged Units KB",
to_char(  
            CASE WHEN  sum(Charged_Units)/1024 >= 77824
            then (sum(Charged_Units)/1024)
         Else 0.00
         End 
   ,
    'fm9999999990.9000'
  ) 
 "Plan B Charged Units KB"
from IOT_AGGREGATOR_ON_NET_USAGE where tadig = '{tadig_code}'
and (ods_insert_date >= to_date('{timeStamp}','YYYYMMDD') 
and ods_insert_date < to_date('{endTimeStamp}','YYYYMMDD')  )
group by
imsi
order by
imsi
"""


headings = {}
 
headings["IoT"] = [
"IMSI",
"Total Records",
"Serving Location",
"Roaming Partner",
"Total Session Duration (Minutes)",
"Incoming KB",
"Outgoing KB",
"Charged Units KB", 
"Total Record Charges $"
];

headings["IoT_IMSI"] = [
"IMSI",
"Total Records",
"Total Session Duration (Minutes)",
"Incoming KB",
"Outgoing KB",
"Total Volume KB",
"Total Record Charges $",
"Plan A Charged Units KB",
"Plan B Charged Units KB"
];


tab = {}
 
tab["IoT"] = "IoT FloLive by Carrier Place";
tab["IoT_IMSI"] = "IoT FloLive by IMSI"; 
tab["IoT_IMSI_USAUF"] = "IoT FloLive by IMSI - USAUF"; 

def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "david.balchen@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("david.balchen@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def dbConnect ():
    
    CONN_INFO = {
        'host': '10.180.171.15',  # info from tnsnames.ora
        'port': 1530,  # info from tnsnames.ora
        'user': 'md1dbal1',
        'psw': 'Potat000#',
        'service': 'bodsdev_adhoc'  # info from tnsnames.ora
    }
    CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)
  
#    tconn = cx_Oracle.connect(CONN_STR)

    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_SVC_BILLINGOPS")
    
    return tconn;


def printRow (row, font, sheet, max_row, max_col=0): 
    for i in range(0, len(row) - max_col):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font
        if i > 2: 
            sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00'             
    return;


def printSheet (title, header, sheet, output, flag=0):
    sheet.title = title
    
    printRow(header, bold_font, sheet, 1)
    max_row = 2 

    for row in output:
        font = def_font

        # Modify fonts here

        printRow(row, font, sheet, max_row, (len(row) - len(header)))
        max_row = max_row + 1 
    return

if __name__ == '__main__':
    pass

timeStamp = sys.argv[1]

endTimeStamp = timeStamp[0:6] + "01"
timeStamp = endTimeStamp

thisMonth = (datetime.strptime(endTimeStamp, '%Y%m%d')).strftime("%B")

timeStamp = datetime.strptime(timeStamp, "%Y%m%d")
timeStamp = (timeStamp - relativedelta(months=1)).strftime('%Y%m%d')

day = "1"
month = timeStamp[4:6]
monthName = (datetime.strptime(timeStamp, '%Y%m%d')).strftime("%B")
year = timeStamp[0:4]

title = """The IoT FloLive  Report for """ + thisMonth + """ """ + day + """, """ + year 
#message = " Attached to this email is the IoT FloLive Report for " + thisMonth + """ """ + day + """, """ + year + "\n\n"
message = "Attached to this email is the IoT FloLive Report for " + thisMonth + """ """ + day + """, """ + year + "\n\n"

# Open an Excel file for output

excel_file = title + '.xlsx'
wb = Workbook()

# Database

conn = dbConnect()
cursor = conn.cursor()
#  
# Do FloLive

results = []
sql = sqlDictionary["IoT"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp)
print(sql)
 
cursor.execute(sql)
results = cursor.fetchall()

# for line in fileinput.input("/home/dbalchen/Desktop/IoTest1.csv"):
#     try:
#         line = line.rstrip()
#         results.append(tuple(line.split("\t")))
#     except:pass

printSheet(tab["IoT"], headings["IoT"], wb.active, results, 1);

 
results = []
sql = sqlDictionary["IoT_IMSI"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp,tadig_code = "IOTFL")
print(sql)

cursor.execute(sql)
results = cursor.fetchall()

# for line in fileinput.input("/home/dbalchen/Desktop/IoTest2.csv"):
#     try:
#         line = line.rstrip()
#         results.append(tuple(line.split("\t")))
#     except:pass

printSheet(tab["IoT_IMSI"], headings["IoT_IMSI"], wb.create_sheet(tab["IoT_IMSI"]), results, 1);


results = []
sql = sqlDictionary["IoT_IMSI"].format(timeStamp=timeStamp, endTimeStamp=endTimeStamp,tadig_code = "USAUF")
print(sql)

cursor.execute(sql)
results = cursor.fetchall()

# for line in fileinput.input("/home/dbalchen/Desktop/IoTest2.csv"):
#     try:
#         line = line.rstrip()
#         results.append(tuple(line.split("\t")))
#     except:pass

printSheet(tab["IoT_IMSI_USAUF"], headings["IoT_IMSI"], wb.create_sheet(tab["IoT_IMSI_USAUF"]), results, 1);


wb.save(excel_file)

# Close database connection

cursor.close()
 
conn.close()


for who in sendTo:
    sendMail(excel_file, message, title, who)
    
SystemExit(0);

