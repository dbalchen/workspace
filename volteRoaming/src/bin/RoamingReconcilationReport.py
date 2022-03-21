#! /usr/bin/python3

'''
Created on Jan 11, 2021
@author: dbalchen

'''
# Oracle Libraries
import cx_Oracle

import fileinput

# libraries used for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# import Workbook used for writing
from openpyxl import Workbook
from openpyxl.styles import Font

from dateutil.relativedelta import relativedelta
from datetime import datetime

import sys
from rrlib import *


def sendMail (xfile, mesg, subject, who):
    
    msg = MIMEMultipart()
    msg['From'] = "david.balchen@uscellular.com"
    msg['To'] = who
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(mesg))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xfile, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="' + xfile + '"')
    msg.attach(part)

    smtp = smtplib.SMTP("localhost", 25)
    smtp.sendmail("david.balchen@uscellular.com", who, msg.as_string())
    smtp.quit()
    return;


def dbConnect ():
    
    CONN_INFO = {
        'host': '10.176.199.19',  # info from tnsnames.ora
        'port': 1530,  # info from tnsnames.ora
        'user': 'md1dbal1',
        'psw': 'Poiu#0987',
        'service': 'bodsprd_adhoc'  # info from tnsnames.ora
    }
    CONN_STR = '{user}/{psw}@{host}:{port}/{service}'.format(**CONN_INFO)
   
#    tconn = cx_Oracle.connect(CONN_STR)

    tconn = cx_Oracle.connect(user='', password='', dsn="BODS_DAV_BILLINGOPS")
    
    return tconn;


def printRow (row, font, sheet, max_row, max_col=0): 
    for i in range(0, len(row) - max_col):
        sheet.cell(row=max_row, column=(i + 1)).value = row[i] 
        sheet.cell(row=max_row, column=(i + 1)).font = font
        if i > 2: 
            sheet.cell(row=max_row, column=(i + 1)).number_format = '0.00'             
    return;


def printSheet (title, header, sheet, output, flag=0):
    sheet.title = title
    
    printRow(header, bold_font, sheet, 1)
    max_row = 2 

    for row in output:
        font = def_font
        if (flag == 1 and row[0] != "") and ((float(row[len(header)]) > 3.00) or (float(row[len(header) + 1]) > 3.00) or (float(row[len(header) + 2]) > 3.00) or (float(row[len(header) + 3]) > 3.00) or (float(row[len(header) + 4]) > 3.00)):
            font = red_font            
        printRow(row, font, sheet, max_row, (len(row) - len(header)))
        max_row = max_row + 1 
    return


def sumColumn (col, rows):
        column_sum = 0
        for crow in rows:
            if crow[col] != '':
                column_sum += float(crow[col])                
        return column_sum


if __name__ == '__main__':
    pass

# sys.argv[1] = "SDIRI_FCIBER,SDATACBR_FDATACBR,CIBER_CIBER,DATA_CIBER,LTE_DATA,LTE_VOLTE,NLDLT,DISP_RM";
sysargv = sys.argv[1]

# sysargv = "SDIRI_FCIBER,SDATACBR_FDATACBR,CIBER_CIBER,DATA_CIBER,LTE,NLDLT,DISP_RM";

timeStamp = sys.argv[2]

day = timeStamp[6:8]
month = timeStamp[4:6]
monthName = (datetime.strptime(timeStamp, '%Y%m%d')).strftime("%B")

year = timeStamp[0:4]

outTimeStamp = datetime.strptime(timeStamp, "%Y%m%d")
outTimeStamp = (outTimeStamp - relativedelta(days=1)).strftime('%Y%m%d')

title = """The Roaming Reconciliation Report for """ + monthName + """ """ + day + """, """ + year 
message = " Attached to this email is the Roaming Reconciliation Report for " + monthName + """ """ + day + """, """ + year + "\n\n"

# Open an Excel file for output

# Font setup   
def_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=False)
bold_font = Font(name='Arial', size=10, color='FF000000', italic=False, bold=True)
red_font = Font(name='Arial', size=10, color='00FF0000', italic=True, bold=False)

excel_file = title + '.xlsx'
wb = Workbook()

# Database
conn = dbConnect()
cursor = conn.cursor()

results = []  # raw results
results2 = []

switches = sysargv.split(',')

for idx, switch in enumerate(switches): 
    print(switch)
    error_sql = '' 
    aprm_sql = ''
    sql = ''

    if (switch == 'LTE') or (switch == 'SDIRI_FCIBER') or (switch == 'SDATACBR_FDATACBR') or (switch == 'NLDLT'):
        sql = sqlDictionary[switch].format(timeStamp=timeStamp)

        error_sql = sqlDictionary["REJECTED_RECORDS"].format(switch=switch, timeStamp=timeStamp)
        
        tmp = switch + "_APRM"        
        aprm_sql = sqlDictionary[tmp].format(timeStamp=timeStamp)
        
    elif (switch == 'DISP_RM') or (switch == 'DATA_CIBER') or (switch == 'CIBER_CIBER') or (switch == 'NLDLT'): 
        sql = sqlDictionary[switch].format(outTimeStamp=outTimeStamp)
        
        tmp = switch + "_APRM"      
        aprm_sql = sqlDictionary[tmp].format(outTimeStamp=outTimeStamp)

    print(sql)

#     for line in fileinput.input("/home/dbalchen/Desktop/test.csv"):
#         try:
#             line = line.rstrip()
#             results.append(tuple(line.split("\t")))
#         except:pass

    cursor.execute(sql)
    results = cursor.fetchall()

    if len(results) == 0:
        continue
    
    print (aprm_sql)
    
#     for line in fileinput.input("/home/dbalchen/Desktop/testA.csv"):
#         try:
#             line = line.rstrip()
#             results2.append(tuple(line.split("\t")))
#         except:pass
    
    cursor.execute(aprm_sql)
    results2 = cursor.fetchall()

    anomalies = [x for x in results if ((float(x[len(headings[switch])]) >= 3.00) 
        or (float(x[len(headings[switch]) + 1]) >= 3.00) 
        or (float(x[len(headings[switch]) + 2]) >= 3.00) 
        or (float(x[len(headings[switch]) + 3]) >= 3.00) 
        or (float(x[len(headings[switch]) + 4]) >= 3.00))
        ]
    
    if len(anomalies) > 0:
        message = message + "\n " + switch + "\n"
        
        for row in anomalies:
            if float(row[len(headings[switch])]) > 3.00:
                message = message + "\n    " + str(row[0]) + ": -  DCH record discrepancy " + "{:.2f}".format(float(row[len(headings[switch])])) + '%' + "\n"

            if float(row[len(headings[switch]) + 1]) > 3.00:
                message = message + "\n    " + str(row[0]) + ": - DCH Volume  discrepancy " + "{:.2f}".format(float(row[len(headings[switch]) + 1])) + '%' + "\n"

            if float(row[len(headings[switch]) + 2]) > 3.00:
                message = message + "\n    " + str(row[0]) + ": -  DCH Charges  discrepancy " + "{:.2f}".format(float(row[len(headings[switch]) + 2])) + '%' + "\n"

            if float(row[len(headings[switch]) + 3]) > 3.00:
              message = message + "\n    " + str(row[0]) + ": -  TC/APRM record discrepancy " + "{:.2f}".format(float(row[len(headings[switch]) + 3])) + '%' + "\n"
            
            if float(row[len(headings[switch]) + 4]) > 3.00:
                message = message + "\n    " + str(row[0]) + ": -  TC/APRM charges discrepancy " + "{:.2f}".format(float(row[len(headings[switch]) + 4])) + '%' + "\n"
    
        message = message + "\n   Usage Type Summary - All Files :\n"
                
        column_sum = sumColumn(len(headings[switch]), results) / float(len(results))
        message = message + "\n    Total Record Difference    :  " + "{:.2f}".format(column_sum) + '% ' + "\n"
            
        column_sum = sumColumn(len(headings[switch]) + 1, results) / float(len(results))       
        message = message + "\n    Total Volume Difference    :  " + "{:.2f}".format(column_sum) + '% ' + "\n"

        column_sum = sumColumn(len(headings[switch]) + 2, results) / float(len(results))
        message = message + "\n    Total Charge Difference    :  " + "{:.2f}".format(column_sum) + '% ' + "\n"
    
    brow = [""] * (len(headings[switch]) + 5)
    
    results.append(tuple(brow))

    sumRow = [""] * (len(headings[switch]) + 5)  
    
    if switch == 'SDIRI_FCIBER':
        sumRow[3] = sumColumn(3, results)
        sumRow[6] = sumColumn(6, results)
        sumRow[4] = sumColumn(4, results)
        sumRow[7] = sumColumn(7, results)
        sumRow[21] = sumColumn(21, results)
        
        results.append(tuple(sumRow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        
        brow[6] = sumRow[6] - sumRow[3]
        brow[7] = sumRow[7] - sumRow[4]
        brow[21] = sumRow[4] - sumRow[21]
        
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        
        crow[6] = brow[6] / sumRow[6]
        crow[7] = brow[7] / sumRow[7]
        crow[21] = brow[21] / sumRow[21]
        
        results.append(tuple(crow))
        
        brow = [""] * (len(headings[switch]) + 5)    
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5)   
        aprmVolume = sumColumn(3, results2)
        
        brow[21] = aprmVolume
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[21] = sumRow[3] - aprmVolume;
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[21] = sumRow[3] / aprmVolume;
        results.append(tuple(brow))

    if switch == 'SDATACBR_FDATACBR':
        sumRow[5] = sumColumn(5, results)
        sumRow[6] = sumColumn(6, results)
        sumRow[10] = sumColumn(10, results)
        sumRow[11] = sumColumn(11, results)
        sumRow[25] = sumColumn(25, results)

        results.append(tuple(sumRow))
        brow = [""] * (len(headings[switch]) + 5) 
        
        brow[10] = sumRow[5] - sumRow[10]
        brow[11] = sumRow[6] - sumRow[11]
        brow[25] = sumRow[6] - sumRow[25]
        
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        
        crow[10] = brow[10] / sumRow[10]
        crow[11] = brow[11] / sumRow[11]
        crow[25] = brow[25] / sumRow[25]
        
        results.append(tuple(crow))   
        
        brow = [""] * (len(headings[switch]) + 5)    
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5)   
        aprmVolume = sumColumn(5, results2)
        
        brow[25] = aprmVolume
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[25] = sumRow[5] - aprmVolume;
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[25] = sumRow[5] / aprmVolume;
        results.append(tuple(brow))    
        
    if switch == 'CIBER_CIBER':
        sumRow[3] = sumColumn(3, results)
        sumRow[5] = sumColumn(5, results)
        sumRow[6] = sumColumn(6, results)
        sumRow[10] = sumColumn(10, results)
        sumRow[11] = sumColumn(11, results)
        
        results.append(tuple(sumRow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[3] = sumRow[3] - sumRow[11]
        brow[5] = sumRow[5] - sumRow[10]
        brow[6] = sumRow[6] - sumRow[11]
        
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        crow[3] = brow[3] / sumRow[3]
        crow[5] = brow[5] / sumRow[5]
        crow[6] = brow[6] / sumRow[6]
        results.append(tuple(crow))
                        
        brow = [""] * (len(headings[switch]) + 5)    
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5)   
        aprmVolume = sumColumn(4, results2)
        
        brow[3] = aprmVolume
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[3] = sumRow[10] - aprmVolume;
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[3] = sumRow[10] / aprmVolume;
        results.append(tuple(brow))    
        
    if switch == 'DATA_CIBER':
        sumRow[2] = sumColumn(2, results)
        sumRow[4] = sumColumn(4, results)
        sumRow[5] = sumColumn(5, results)
        sumRow[9] = sumColumn(9, results)
        sumRow[10] = sumColumn(10, results)
        
        results.append(tuple(sumRow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        
        brow[9] = sumRow[4] - sumRow[9]
        brow[10] = sumRow[5] - sumRow[10]
 
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        
        crow[9] = brow[9] / sumRow[9]
        crow[10] = brow[10] / sumRow[10]
        crow[9] = brow[9] / sumRow[9]
        crow[10] = brow[10] / sumRow[10]       
        results.append(tuple(crow))

    if switch == 'LTE':
        sumRow[6] = sumColumn(6, results)
        sumRow[7] = sumColumn(7, results)
        sumRow[11] = sumColumn(11, results) 
        sumRow[12] = sumColumn(12, results)
        sumRow[21] = sumColumn(21, results)
        sumRow[26] = sumColumn(26, results) 
        sumRow[29] = sumColumn(29, results)
        
        results.append(tuple(sumRow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        
        brow[11] = sumRow[11] - sumRow[6]
        brow[12] = sumRow[12] - sumRow[7]
        brow[21] = sumRow[21] - sumRow[7]
        
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        
        crow[11] = brow[11] / sumRow[11]
        crow[12] = brow[12] / sumRow[12]
        crow[21] = brow[21] / sumRow[21]
        
        results.append(tuple(crow))
               
        brow = [""] * (len(headings[switch]) + 5)    
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5)   
        aprmVolume = sumColumn(5, results2)
        
        brow[21] = aprmVolume
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[21] = sumRow[6] - aprmVolume;
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[21] = sumRow[6] / aprmVolume;
        results.append(tuple(brow))    
        
    if switch == 'NLDLT':
        sumRow[5] = sumColumn(5, results)
        sumRow[6] = sumColumn(6, results)
        sumRow[8] = sumColumn(8, results)
        sumRow[9] = sumColumn(9, results) 
        sumRow[18] = sumColumn(18, results)   
        
        results.append(tuple(sumRow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        
        brow[8] = sumRow[8] - sumRow[5]
        brow[9] = sumRow[9] - sumRow[6]
        brow[18] = sumRow[18] - sumRow[6]
        
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        
        crow[8] = brow[8] / sumRow[8]
        crow[9] = brow[9] / sumRow[9]
        crow[18] = brow[18] / sumRow[18]
        
        results.append(tuple(crow))   
        
        brow = [""] * (len(headings[switch]) + 5)    
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5)   
        aprmVolume = sumColumn(5, results2)
        
        brow[18] = aprmVolume
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[18] = sumRow[5] - aprmVolume;
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[18] = sumRow[5] / aprmVolume;
        results.append(tuple(brow))             
          
    if switch == 'DISP_RM':
        sumRow[2] = sumColumn(2, results)
        sumRow[3] = sumColumn(3, results)
        sumRow[5] = sumColumn(8, results)
        sumRow[6] = sumColumn(9, results)
        sumRow[8] = sumColumn(11, results) 
        sumRow[9] = sumColumn(9, results)  
        sumRow[11] = sumColumn(11, results)
        sumRow[12] = sumColumn(12, results)
        results.append(tuple(sumRow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        
        brow[8] = sumRow[8] - sumRow[11]
        brow[9] = sumRow[9] - sumRow[12]
        
        results.append(tuple(brow))
        
        crow = [""] * (len(headings[switch]) + 5) 
        
        crow[8] = brow[8] / sumRow[8]
        crow[9] = brow[9] / sumRow[9]
        
        results.append(tuple(crow)) 
                
        brow = [""] * (len(headings[switch]) + 5)    
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5)  
        aprmVolume = sumColumn(4, results2)
        aprmVolume2 = sumColumn(5, results2)
        
        brow[11] = aprmVolume
        brow[12] = aprmVolume2
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[11] = sumRow[11] - aprmVolume;
        brow[12] = sumRow[12] - aprmVolume2;
        results.append(tuple(brow))
        
        brow = [""] * (len(headings[switch]) + 5) 
        brow[11] = sumRow[11] / aprmVolume;
        brow[12] = sumRow[12] / aprmVolume2;
        
        results.append(tuple(brow))    
                  
    if idx == 0:
        printSheet(tab[switch], headings[switch], wb.active, results, 1)        
    else:
        printSheet(tab[switch], headings[switch], wb.create_sheet(tab[switch]), results, 1);

    if error_sql != '' :
        print(error_sql)
        
        cursor.execute(error_sql)
        results = cursor.fetchall()

        tap = tab[switch] + " Errors"
        printSheet(tap, headings['ERROR'], wb.create_sheet(tap), results);

    tap = tab[switch] + "_APRM"
    tap2 = switch + "_APRM"

    printSheet(tap, headings[tap2], wb.create_sheet(tap), results2);
    
wb.save(excel_file)

# Close database connection
cursor.close()
conn.close()

#sendTo = ["david.balchen@uscellular.com"]
sendTo = ["david.balchen@uscellular.com", 'Kevin.Hergenrother@uscellular.com','ISBillingOperations@uscellular.com', 'Ilham.Elgarni@uscellular.com', 'david.smith@uscellular.com']

for who in sendTo:
    sendMail(excel_file, message, title, who)
    
SystemExit(0);
